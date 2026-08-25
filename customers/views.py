from io import BytesIO

from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from billing.models import Bill
from billing.services.bill_service import BillService
from pdf.services.pdf_service import PDFService
from reviews.services.review_tracking_service import ReviewTrackingService

from .forms import BillForm, CustomerForm
from .models import Customer
from .services.customer_service import CustomerService


SORT_FIELDS = {
    'name': 'name',
    'phone': 'phone',
    'email': 'email',
    'bill_amount': 'bill_amount',
    'table_number': 'table_number',
    'visit_date': 'visit_date',
    'created_at': 'created_at',
}


def dashboard(request):
    stats = CustomerService.dashboard_stats()
    stats['review_clicks'] = ReviewTrackingService.count_clicks()
    recent_customers = Customer.objects.all()[:8]
    return render(
        request,
        'customers/dashboard.html',
        {
            'stats': stats,
            'recent_customers': recent_customers,
            'form': CustomerForm(),
        },
    )


def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = CustomerService.save_customer_and_notify(form)
            messages.success(request, f'{customer.name} was saved. Email and WhatsApp notifications are being processed.')
            return redirect('customers:detail', pk=customer.pk)
        messages.error(request, 'Please correct the highlighted fields.')
    else:
        form = CustomerForm()

    return render(request, 'customers/customer_form.html', {'form': form, 'title': 'Add Customer'})


def customer_list(request):
    customers = _filtered_customers(request)
    paginator = Paginator(customers, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(
        request,
        'customers/customer_list.html',
        {
            'page_obj': page_obj,
            'search': request.GET.get('search', ''),
            'status': request.GET.get('status', ''),
            'sort': request.GET.get('sort', '-visit_date'),
        },
    )


def customer_detail(request, pk):
    customer = get_object_or_404(Customer.objects.prefetch_related('bills__items', 'notification_logs'), pk=pk)
    latest_bill = customer.bills.first()
    latest_notification = customer.notification_logs.first()
    return render(
        request,
        'customers/customer_detail.html',
        {
            'customer': customer,
            'latest_bill': latest_bill,
            'bills': customer.bills.all(),
            'latest_notification': latest_notification,
        },
    )


def customer_update(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'{customer.name} was updated successfully.')
            return redirect('customers:detail', pk=customer.pk)
        messages.error(request, 'Please correct the highlighted fields.')
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'customers/customer_form.html', {'form': form, 'title': 'Edit Customer', 'customer': customer})


def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer_name = customer.name
        customer.delete()
        messages.success(request, f'{customer_name} was deleted successfully.')
        return redirect('customers:list')
    return render(request, 'customers/customer_confirm_delete.html', {'customer': customer})


def bill_create(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = BillForm(request.POST)
        if form.is_valid():
            bill = CustomerService.add_bill_and_notify(customer, form)
            messages.success(request, f'New bill {bill.bill_number} was added. WhatsApp/email notifications are being processed for this bill.')
            return redirect('customers:detail', pk=customer.pk)
        messages.error(request, 'Please correct the highlighted bill fields.')
    else:
        form = BillForm(initial={
            'table_number': customer.table_number,
            'visit_date': customer.visit_date,
        })
    return render(request, 'customers/bill_form.html', {'form': form, 'customer': customer, 'title': 'Add New Bill'})


def bill_update(request, pk, bill_id):
    customer = get_object_or_404(Customer, pk=pk)
    bill = get_object_or_404(Bill.objects.prefetch_related('items'), pk=bill_id, customer=customer)
    if request.method == 'POST':
        form = BillForm(request.POST, bill=bill)
        if form.is_valid():
            bill = BillService.update_bill(bill, form.cleaned_data)
            PDFService.generate_bill_pdf(bill)
            CustomerService.sync_customer_latest_bill(customer)
            messages.success(request, f'Bill {bill.bill_number} was updated. Use Resend WhatsApp if you want to send the updated bill again.')
            return redirect('customers:detail', pk=customer.pk)
        messages.error(request, 'Please correct the highlighted bill fields.')
    else:
        form = BillForm(bill=bill)
    return render(request, 'customers/bill_form.html', {'form': form, 'customer': customer, 'bill': bill, 'title': 'Edit Bill'})


def bill_delete(request, pk, bill_id):
    customer = get_object_or_404(Customer, pk=pk)
    bill = get_object_or_404(Bill, pk=bill_id, customer=customer)
    if request.method == 'POST':
        bill_number = bill.bill_number
        if bill.pdf_file:
            bill.pdf_file.delete(save=False)
        bill.delete()
        CustomerService.sync_customer_latest_bill(customer)
        messages.success(request, f'Bill {bill_number} was deleted.')
        return redirect('customers:detail', pk=customer.pk)
    return render(request, 'customers/bill_confirm_delete.html', {'customer': customer, 'bill': bill})


def export_customers_excel(request):
    customers = _filtered_customers(request)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Food Lab Customers'

    headers = [
        'Name',
        'Phone',
        'Email',
        'Bill Amount',
        'Table Number',
        'Visit Date',
        'Notes',
        'Email Sent',
        'WhatsApp Sent',
        'Latest Bill Number',
        'Created At',
    ]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='8B1E3F')

    for customer in customers:
        worksheet.append([
            customer.name,
            customer.phone,
            customer.email,
            float(customer.bill_amount) if customer.bill_amount is not None else '',
            customer.table_number,
            customer.visit_date.isoformat(),
            customer.notes,
            'Yes' if customer.email_sent else 'No',
            'Yes' if customer.whatsapp_sent else 'No',
            customer.bills.first().bill_number if customer.bills.exists() else '',
            customer.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    for column_cells in worksheet.columns:
        width = max(len(str(cell.value or '')) for cell in column_cells) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(width, 45)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="food-lab-customers.xlsx"'
    return response


def download_bill(request, bill_id):
    bill = get_object_or_404(Bill, pk=bill_id)
    if not bill.pdf_file:
        raise Http404('Bill PDF is not available.')
    return FileResponse(open(bill.pdf_file.path, 'rb'), as_attachment=True, filename=f'food-lab-bill-{bill.bill_number}.pdf')


def resend_bill(request, pk, bill_id):
    if request.method != 'POST':
        return redirect('customers:detail', pk=pk)
    customer = get_object_or_404(Customer, pk=pk)
    bill = get_object_or_404(Bill, pk=bill_id, customer=customer)
    result = CustomerService.resend_bill(customer, bill)
    if result['sent']:
        messages.success(request, 'Bill PDF was resent on WhatsApp.')
    else:
        messages.error(request, 'Bill PDF could not be resent on WhatsApp. Check notification logs.')
    return redirect('customers:detail', pk=customer.pk)


def resend_whatsapp(request, pk, bill_id):
    if request.method != 'POST':
        return redirect('customers:detail', pk=pk)
    customer = get_object_or_404(Customer, pk=pk)
    bill = get_object_or_404(Bill, pk=bill_id, customer=customer)
    result = CustomerService.resend_whatsapp(customer, bill)
    if result['sent']:
        messages.success(request, 'WhatsApp bill and thank-you message were resent.')
    else:
        messages.error(request, 'WhatsApp resend failed. Check notification logs.')
    return redirect('customers:detail', pk=customer.pk)


def resend_email(request, pk, bill_id):
    if request.method != 'POST':
        return redirect('customers:detail', pk=pk)
    customer = get_object_or_404(Customer, pk=pk)
    bill = get_object_or_404(Bill, pk=bill_id, customer=customer)
    result = CustomerService.resend_email(customer, bill)
    if result['sent']:
        messages.success(request, 'Thank-you email was resent.')
    else:
        messages.error(request, 'Email resend failed. Check SMTP settings and notification logs.')
    return redirect('customers:detail', pk=customer.pk)


def _filtered_customers(request):
    queryset = Customer.objects.all()
    search = request.GET.get('search', '').strip()
    status = request.GET.get('status', '').strip()
    sort = request.GET.get('sort', '-visit_date')

    if search:
        queryset = queryset.filter(
            Q(name__icontains=search)
            | Q(phone__icontains=search)
            | Q(email__icontains=search)
            | Q(table_number__icontains=search)
        )

    if status == 'email_sent':
        queryset = queryset.filter(email_sent=True)
    elif status == 'whatsapp_sent':
        queryset = queryset.filter(whatsapp_sent=True)
    elif status == 'pending_email':
        queryset = queryset.filter(email__gt='', email_sent=False)
    elif status == 'pending_whatsapp':
        queryset = queryset.filter(whatsapp_sent=False)

    descending = sort.startswith('-')
    field_name = sort[1:] if descending else sort
    sort_field = SORT_FIELDS.get(field_name, 'visit_date')
    if descending:
        sort_field = f'-{sort_field}'
    return queryset.order_by(sort_field, '-created_at')
